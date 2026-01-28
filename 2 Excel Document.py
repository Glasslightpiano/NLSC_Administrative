import pyodbc
import pandas as pd
import openpyxl
import warnings  # 與腳本執行無關，單純隱藏UserWarning(使用者警告)
import arcpy
import os
import glob
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
warnings.filterwarnings("ignore", category=UserWarning)  # 隱藏所有 UserWarning

# ==============================================================================
# 【人工手動輸入參數區】 - 請在此修改執行條件
# ==============================================================================

# --- 修正清單 參數 ---
mdb_file = r'D:\行政界線文件匯出\行政界線_2026-01-19.mdb'
output_path = r'D:\行政界線文件匯出\修正清單_1150119.xlsx'
village_date_str = "2026-01-14"

# --- 面積清冊 參數 ---
TEST_MODE = "SHP"  # 資料來源模式: "SHP" (Shapefile), "GDB_T" (GDB topology check), "GDB_M" (from MDB Query select)
outExcel = f"面積差異分析表_1150119.xlsx"

# --- 參照範本 ---
template_path_1 = r'D:\行政界線文件匯出\母本檔案區\修正清單_空白範本.xlsx'
template_path_2 = r'D:\行政界線文件匯出\母本檔案區\面積差異分析表_空白範本.xlsx'  # 確保此範本無合併儲存格

# ==============================================================================
# 第一階段：執行 匯出修正清單
# ==============================================================================
print(">>> 正在執行第一階段：匯出修正清單...")

try:
    # 建立連線並讀取 MDB 資料
    conn = pyodbc.connect(r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + mdb_file + ';')
    df_village = pd.read_sql("SELECT * FROM Village_NLSC", conn)
    df_modify = pd.read_sql("SELECT * FROM Modify_Data", conn)
    conn.close()

    # 篩選 Village_NLSC 資料
    t_date_v = pd.to_datetime(village_date_str)
    df_village['Add_Date_dt'] = pd.to_datetime(df_village['Add_Date'], errors='coerce')
    df_village['Del_Date_dt'] = pd.to_datetime(df_village['Del_Date'], errors='coerce')

    filtered_v = df_village[
        (df_village['Add_Date_dt'].dt.date == t_date_v.date()) |
        (df_village['Del_Date_dt'].dt.date == t_date_v.date())
        ].copy()
    # Village_NLSC 重複值處理與排序
    if not filtered_v.empty:
        before_count = len(filtered_v)
        filtered_v = filtered_v.drop_duplicates(subset=['VILLAGE_ID'], keep='first')
        filtered_v = filtered_v.sort_values(by='VILLAGE_ID', ascending=True)  # 村里代碼由小排到大
        print(f"成功處理！原始符合 {before_count} 筆，去重並排序後保留 {len(filtered_v)} 筆。")
    else:
        print(f"Village_NLSC 篩選日期 {village_date_str}：查無資料。")

    # 處理 Modify_Data 並建立 CASE_ID 對應表
    t_date_m = pd.to_datetime(village_date_str)
    # 格式化日期為 YYYY/MM/DD 供 K 欄使用
    formatted_m_date = t_date_m.strftime('%Y/%m/%d')

    df_modify['M_Date_dt'] = pd.to_datetime(df_modify['M_Date'], errors='coerce')
    filtered_m = df_modify[df_modify['M_Date_dt'].dt.date == t_date_m.date()].copy()

    # 建立一個字典來存儲 VILLAGE_ID 與 CASE_ID 的關係
    case_mapping = {}
    for _, row in filtered_m.iterrows():
        case_id = str(row['CASE_ID'])
        # 拆分 Admin_ID (以分號拆開)
        admin_ids = str(row['Admin_ID']).split(';')
        for vid in admin_ids:
            vid = vid.strip()
            if vid not in case_mapping:
                case_mapping[vid] = []
            if case_id not in case_mapping[vid]:   # 避免重複加入
                case_mapping[vid].append(case_id)

    # 整合資料並寫入範本
    if not filtered_v.empty:
        wb1 = openpyxl.load_workbook(template_path_1)
        ws1 = wb1['村里'] if '村里' in wb1.sheetnames else wb1.active

        # 定義 Excel 字體、對齊、框線樣式
        my_font = Font(name='微軟正黑體', size=12)  # 字型、字體大小
        my_align = Alignment(horizontal='center', vertical='center')  # 左右置中、上下置中
        thin_side = Side(border_style="thin", color="000000")  # 黑色細線
        my_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)  # 儲存格4邊框線

        current_row = 3
        for _, v_row in filtered_v.iterrows():
            # A-D 欄 (C_Name ~ T_Name_e)
            ws1.cell(row=current_row, column=1, value=v_row.get('C_Name'))
            ws1.cell(row=current_row, column=2, value=v_row.get('C_Name_e'))
            ws1.cell(row=current_row, column=3, value=v_row.get('T_Name'))
            ws1.cell(row=current_row, column=4, value=v_row.get('T_Name_e'))

            # E 欄 (Town_ID)
            ws1.cell(row=current_row, column=5, value=v_row.get('TOWN_ID'))

            # F-H 欄 (V_Name ~ VILLAGE_ID)
            ws1.cell(row=current_row, column=6, value=v_row.get('V_Name'))
            ws1.cell(row=current_row, column=7, value=v_row.get('V_Name_e'))
            ws1.cell(row=current_row, column=8, value=v_row.get('VILLAGE_ID'))

            # I 欄 (CASE_ID 比對)
            vid = str(v_row.get('VILLAGE_ID'))
            case_val = "、".join(case_mapping[vid]) if vid in case_mapping else "-"
            ws1.cell(row=current_row, column=9, value=case_val)

            # K 欄 (修正日期)
            ws1.cell(row=current_row, column=11, value=formatted_m_date)

            # --- 套用樣式到該列的所有有資料欄位 (A-L 欄) ---
            for col in range(1, 13):
                cell = ws1.cell(row=current_row, column=col)
                # cell.font = my_font  # 基礎寫法
                # cell.alignment = my_align
                # cell.border = my_border
                cell.font, cell.alignment, cell.border = my_font, my_align, my_border  # 進階寫法：一行搞定
            current_row += 1

        wb1.save(output_path)
        print(f"第一階段完成：修正清單已儲存。")

except Exception as e:
    print(f"第一階段發生錯誤：{e}")

# ==============================================================================
# 第二階段：執行 多鄉鎮分表匯出，最終才合併
# ==============================================================================
print("\n>>> 正在執行第二階段：面積差異分析表...")

# 讀取目標鄉鎮代碼 (E 欄)
df_temp = pd.read_excel(output_path, usecols="E", skiprows=2, header=None)  # skiprows=2 略過前兩行；header=None 無標題
target_towncodes = df_temp[4].dropna().unique().tolist()  # .dropna() 刪除空值；.unique() 取唯一值；.tolist() 轉換為清單(list)
target_towncodes = [str(int(t)) if isinstance(t, (float, int)) else str(t).strip() for t in target_towncodes]  # 轉文字

# 確認母本對照表與 Excel 範本路徑
if target_towncodes:
    BASE_DIR = r"D:\行政界線文件匯出"
    MASTER_DIR = os.path.join(BASE_DIR, "母本檔案區")
    VILL_MASTER_FILE = os.path.join(MASTER_DIR, "全台村里代碼表.xlsx")
    TOWN_AREA_FILE = os.path.join(MASTER_DIR, "鄉鎮市區面積表.xlsx")
    save_path_2 = os.path.join(BASE_DIR, outExcel)

    # 設定座標系統為 TWD97 / 121分帶 (EPSG:3826)
    TWD97_SR = arcpy.SpatialReference(3826)

    # 定義 Excel 字體、對齊、框線樣式，與修正清單 excel 設定相同
    std_font = Font(name='微軟正黑體', size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(border_style="thin", color="000000")
    std_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 讀取村里對照表並清理字串
    df_vill = pd.read_excel(VILL_MASTER_FILE)
    df_vill['TOWNCODE'] = df_vill['TOWNCODE'].astype(str).str.strip()  # .astype(str)功用：強制轉換字串
    df_vill['VILLCODE'] = df_vill['VILLCODE'].astype(str).str.strip()  # .str.strip()功用：去除字串首尾空格
    # 讀取鄉鎮市區面積表並清理字串
    df_town = pd.read_excel(TOWN_AREA_FILE)
    df_town['TOWNCODE'] = df_town['TOWNCODE'].astype(str).str.strip()

    wb2 = load_workbook(template_path_2)
    ws2 = wb2.active

    # --- 1. 結構預配置：根據鄉鎮數量複製範本區塊 ---
    # 範本結構預設為 1-5 列 (標題x2, 資料預留x1, 總計x1, 空行x1)
    if len(target_towncodes) > 1:  # 多筆 TOWNCODE 才執行
        for i in range(1, len(target_towncodes)):
            dest_start = ws2.max_row + 2  # 間隔 1 列空行
            for r in range(1, 6):
                for c in range(1, 9):
                    source_cell = ws2.cell(row=r, column=c)
                    new_cell = ws2.cell(row=dest_start + r - 1, column=c, value=source_cell.value)
                    if source_cell.has_style:
                        new_cell.font = copy(source_cell.font)
                        new_cell.alignment = copy(source_cell.alignment)
                        new_cell.border = copy(source_cell.border)
                        new_cell.fill = copy(source_cell.fill)
                        new_cell.number_format = source_cell.number_format  # 確保儲存格格式與母本皆相同

    # --- 2. 填充資料 (由上而下處理) ---
    current_block_start = 1  # 標示鄉鎮表格的「起始行號」

    for t_idx, code in enumerate(target_towncodes):  # for迴圈：從最小的 TOWNCODE 開始讀取處理
        print(f"處理鄉鎮資料：{code}")
        # 篩選該鄉鎮市區下的所有村里，並依村里代碼由小排到大
        target_vills = df_vill[df_vill['TOWNCODE'] == code].copy().sort_values('VILLCODE')
        if target_vills.empty:
            continue  # 若村里代碼空值腳本仍會繼續執行

        num_vills = len(target_vills)  # 計算該鄉鎮總共有幾個村里，這會關係到excel的新增欄位數
        #  抓取 TOWNCODE 對應的縣市名稱與鄉鎮市區名稱
        county_name = target_vills.iloc[0]['COUNTYNAME']
        town_name = target_vills.iloc[0]['TOWNNAME']  # .iloc[0] 只需要取篩選結果的第一筆

        # 檢查代碼是否存在於面積表之中
        if code in df_town['TOWNCODE'].values:
            # 存在則取出 Shape_Area 欄位的值
            town_area_series = df_town[df_town['TOWNCODE'] == code]['Shape_Area']  # TOWNCODE對應Shape_Area所在儲存格
            town_area = town_area_series.values[0]  # 取出該儲存格的面積數值
        else:
            # 不存在則預設為 0
            town_area = 0

        # 行政區資訊填寫
        ws2.cell(row=current_block_start + 1, column=1, value=county_name)
        ws2.cell(row=current_block_start + 1, column=2, value=town_name)
        ws2.cell(row=current_block_start + 1, column=3, value=code)
        ws2.cell(row=current_block_start + 1, column=4, value=town_area)
        ws2.cell(row=current_block_start + 1, column=5, value=town_area)

        # 空間路徑判斷
        old_path, new_path = "", ""
        id_field = "VILLCODE" if TEST_MODE == "SHP" else "VILLAGE_ID"  # 三元運算式寫法
        # if TEST_MODE == "SHP":   # 一般 if...else 寫法
        #     id_field = "VILLCODE"
        # else:
        #     id_field = "VILLAGE_ID"

        # SHP 模式：抓取 shp_ 開頭的資料夾中的第一個 shp 檔
        if TEST_MODE == "SHP":
            shp_dirs = sorted(glob.glob(os.path.join(BASE_DIR, "shp_*")))  # 搜尋路徑上名稱為 shp 開頭的資料夾並使用 sorted 排序
            if len(shp_dirs) >= 2:  # 確定有二個資料夾(一個舊，一個新)
                arcpy.env.workspace = shp_dirs[0]  # 處理舊資料(0代表第一筆)，workspace讀取舊資料夾內的全部 shp 相關檔案
                f_o = arcpy.ListFeatureClasses("*.shp")
                if f_o:
                    old_path = os.path.join(shp_dirs[0], f_o[0])  # 如有shp檔，抓第一個（f_o[0]），並合併成完整的案路徑給old_path
                arcpy.env.workspace = shp_dirs[-1]  # 處理新資料(-1代表最後一筆)，workspace讀取舊資料夾內的全部 shp 相關檔案
                f_n = arcpy.ListFeatureClasses("*.shp")
                if f_n:
                    new_path = os.path.join(shp_dirs[-1], f_n[0])  # 如有shp檔，抓第一個（f_n[0]），並合併成完整的案路徑給new_path

        # GDB_T 模式：抓取不同 GDB 內固定路徑下的 Village 圖層
        elif TEST_MODE == "GDB_T":
            gdb_dirs = sorted(glob.glob(os.path.join(BASE_DIR, "administrative_boundary*.gdb")))
            if len(gdb_dirs) >= 2:
                old_path = os.path.join(gdb_dirs[0], "Check_topo_geo", "Village")
                new_path = os.path.join(gdb_dirs[-1], "Check_topo_geo", "Village")

        # GDB_M 模式：單一 GDB 內抓取不同版本的 Village 圖層
        elif TEST_MODE == "GDB_M":
            gdb_m_path = os.path.join(BASE_DIR, "GDB_M.gdb")
            if os.path.exists(gdb_m_path):
                arcpy.env.workspace = gdb_m_path
                layers = sorted(arcpy.ListFeatureClasses("Village_NLSC*"))
                if len(layers) >= 2:
                    old_path = os.path.join(gdb_m_path, layers[0])
                    new_path = os.path.join(gdb_m_path, layers[-1])

        # 定義搜尋與計算函數：使用 SearchCursor 取得各里面積
        def get_area_dict(fileMode, fld, vCodeList):
            d = {}  # 建立 python 字典
            if not fileMode or not arcpy.Exists(fileMode):
                return d
            id_str = ",".join(["'{}'".format(v) for v in vCodeList])  # 為了轉成SQL語法 IN 的使用採單引號
            # 組合 SQL 過濾字串 (例如: VILLCODE IN ('67000190001', '67000190002'))
            where = "{} IN ({})".format(fld, id_str)
            with arcpy.da.SearchCursor(fileMode, [fld, "SHAPE@AREA"], where_clause=where, spatial_reference=TWD97_SR) as cur:
                for rowField in cur:
                    d[str(rowField[0])] = rowField[1]
                # for r in cur: d[str(r[0])] = round(r[1], 4)  # 開會討論後不需要四捨五入到小數點後4位
            return d


        v_codes = target_vills['VILLCODE'].tolist()

        # 計算取得更新前面積(old_areas)與更新後面積(new_areas)
        old_areas = get_area_dict(old_path, id_field, v_codes)
        new_areas = get_area_dict(new_path, id_field, v_codes)

        # 插入行數並填資料
        data_row_start = current_block_start + 3  # 第一筆村里資料填在空白的第4行
        if num_vills > 1:
            # 母本原有一行，所以欲新增行數為該鄉鎮市區村里數-1
            ws2.insert_rows(data_row_start + 1, amount=num_vills - 1)

        sum_o, sum_n = 0.0, 0.0
        # _是 Pandas 內建的列索引，註記底線表示忽略；v_row 紀錄VILLCODE、VILLNAME；enumerate() 計算執行「次數」
        for i, (_, v_row) in enumerate(target_vills.iterrows()):
            r = data_row_start + i
            vc, vn = str(v_row['VILLCODE']), str(v_row['VILLNAME'])  # 確保村里代碼、村里名稱為字串
            # 讀取更新前、更新後面積數值。.get(vc, 0.0)保險作用：若該代碼在圖層中找不到(例如沒這份資料)，則回傳 0.0
            ao, an = old_areas.get(vc, 0.0), new_areas.get(vc, 0.0)
            # 寫入 excel
            ws2.cell(row=r, column=1, value=town_name)
            ws2.cell(row=r, column=2, value=vc)
            ws2.cell(row=r, column=3, value=vn)
            ws2.cell(row=r, column=4, value=ao)
            ws2.cell(row=r, column=5, value=an)
            ws2.cell(row=r, column=6, value=f"=E{r}-D{r}")

            # 儲存格外觀設定套用
            for col in range(1, 9):
                cell = ws2.cell(row=r, column=col)
                cell.font, cell.alignment, cell.border = std_font, center_align, std_border
                if col in [6]:
                    cell.number_format = '0.00'  # 面積差異欄位儲存格格式為數值、小數點後2位

            sum_o += ao  # 更新前面積加總
            sum_n += an  # 更新後面積加總

        # 填寫總計資料(尚未合併儲存格)
        total_row = data_row_start + num_vills
        ws2.cell(row=total_row, column=1, value=f"{town_name} 總計")
        ws2.cell(row=total_row, column=4, value=sum_o)
        ws2.cell(row=total_row, column=5, value=sum_n)
        ws2.cell(row=total_row, column=6, value=f"=E{total_row}-D{total_row}")
        for col in range(1, 9):
            cell = ws2.cell(row=total_row, column=col)
            cell.font, cell.alignment, cell.border = std_font, center_align, std_border
            if col in [4, 5, 6]:
                cell.number_format = '0.00'

        # 更新下一個區塊的起點 (目前總計列 + 1 列空行)
        current_block_start = total_row + 2

    # --- 3. 最終修飾：掃描 A 欄合併所有「總計」列 ---
    print("正在執行最後合併與格式優化...")
    for row in range(1, ws2.max_row + 1):
        cell_val = ws2.cell(row=row, column=1).value
        if cell_val and "總計" in str(cell_val):
            # 將該列的 A, B, C 欄合併
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)

    # 存檔
    wb2.save(save_path_2)
    print(f"\n>>> 任務圓滿完成！檔案路徑：{save_path_2}")
