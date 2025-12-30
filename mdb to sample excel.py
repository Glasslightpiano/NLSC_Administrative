import pyodbc
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

# 1. 設定路徑與參數
mdb_file = r'D:\行政界線文件匯出\行政界線_2025-10-31_test.mdb'
output_path = r'D:\行政界線文件匯出\修正清單_.xlsx'

# --- 日期篩選參數設定 ---
village_date_str = "2024-12-30"
modify_date_str = "2025-10-31"

# --- 空白修正清單範本路徑 ---
template_path = r'D:\行政界線文件匯出\母本檔案區\修正清單_空白範本.xlsx'

try:
    # 2. 建立連線並讀取 MDB 資料
    conn = pyodbc.connect(r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};DBQ=' + mdb_file + ';')
    df_village = pd.read_sql("SELECT * FROM Village_NLSC", conn)
    df_modify = pd.read_sql("SELECT * FROM Modify_Data", conn)
    conn.close()

    # 3. 篩選 Village_NLSC 資料
    t_date_v = pd.to_datetime(village_date_str)
    df_village['Add_Date_dt'] = pd.to_datetime(df_village['Add_Date'], errors='coerce')
    df_village['Del_Date_dt'] = pd.to_datetime(df_village['Del_Date'], errors='coerce')

    filtered_v = df_village[
        (df_village['Add_Date_dt'].dt.date == t_date_v.date()) |
        (df_village['Del_Date_dt'].dt.date == t_date_v.date())
        ].copy()

    # 4. Village_NLSC 重複值處理與排序
    if not filtered_v.empty:
        before_count = len(filtered_v)
        filtered_v = filtered_v.drop_duplicates(subset=['VILLAGE_ID'], keep='first')
        filtered_v = filtered_v.sort_values(by='VILLAGE_ID', ascending=True)
        print(f"成功處理！原始符合 {before_count} 筆，去重並排序後保留 {len(filtered_v)} 筆。")
    else:
        print(f"Village_NLSC 篩選日期 {village_date_str}：查無資料。")

    # 5. 處理 Modify_Data 並建立 CASE_ID 對應表
    t_date_m = pd.to_datetime(modify_date_str)
    # 格式化日期為 YYYY/MM/DD 供 K 欄使用
    formatted_m_date = t_date_m.strftime('%Y/%m/%d')

    df_modify['M_Date_dt'] = pd.to_datetime(df_modify['M_Date'], errors='coerce')
    filtered_m = df_modify[df_modify['M_Date_dt'].dt.date == t_date_m.date()].copy()

    # 建立一個字典來存儲 VILLAGE_ID 與 CASE_ID 的關係
    case_mapping = {}
    for _, row in filtered_m.iterrows():
        case_id = str(row['CASE_ID'])
        # 拆分 Admin_ID (以分號拆開)
        admin_ids = str(row['Admin_ID']).split(';')
        for vid in admin_ids:
            vid = vid.strip()
            if vid not in case_mapping:
                case_mapping[vid] = []
            if case_id not in case_mapping[vid]:  # 避免重複加入
                case_mapping[vid].append(case_id)

    # 6. 整合資料並寫入範本
    if not filtered_v.empty:
        wb = openpyxl.load_workbook(template_path)
        ws = wb['村里'] if '村里' in wb.sheetnames else wb.active

        # --- 定義樣式 (需求設定) ---
        # 1. 字體
        my_font = Font(name='微軟正黑體', size=12)
        # 2. 置中
        my_align = Alignment(horizontal='center', vertical='center')
        # 3. 框線 (黑色細線)
        thin_side = Side(border_style="thin", color="000000")
        my_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        current_row = 3
        for _, v_row in filtered_v.iterrows():
            # A-D 欄 (C_Name ~ T_Name_e)
            ws.cell(row=current_row, column=1, value=v_row.get('C_Name'))
            ws.cell(row=current_row, column=2, value=v_row.get('C_Name_e'))
            ws.cell(row=current_row, column=3, value=v_row.get('T_Name'))
            ws.cell(row=current_row, column=4, value=v_row.get('T_Name_e'))

            # E 欄 (Town_ID)
            ws.cell(row=current_row, column=5, value=v_row.get('TOWN_ID'))

            # F-H 欄 (V_Name ~ VILLAGE_ID)
            ws.cell(row=current_row, column=6, value=v_row.get('V_Name'))
            ws.cell(row=current_row, column=7, value=v_row.get('V_Name_e'))
            ws.cell(row=current_row, column=8, value=v_row.get('VILLAGE_ID'))

            # I 欄 (CASE_ID 比對)
            vid = str(v_row.get('VILLAGE_ID'))
            case_val = "、".join(case_mapping[vid]) if vid in case_mapping else "-"
            ws.cell(row=current_row, column=9, value=case_val)

            # K 欄 (修正日期)
            ws.cell(row=current_row, column=11, value=formatted_m_date)

            # --- 套用樣式到該列的所有有資料欄位 (A-L 欄) ---
            for col in range(1, 13):
                target_cell = ws.cell(row=current_row, column=col)
                target_cell.font = my_font
                target_cell.alignment = my_align
                target_cell.border = my_border  # 加入框線

            current_row += 1

        wb.save(output_path)
        print(f"成功！已套用範本、格式化並加入框線，資料已匯出至 {output_path}")

except Exception as e:
    print(f"發生錯誤：{e}")
